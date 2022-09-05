# 네이버 환율 정보
import urllib.request as req
from bs4 import BeautifulSoup as bs
# url = "https://finance.naver.com/marketindex/exchangeList.naver"
# data = req.urlopen( url ).read().decode( "euc-kr" )
# req.urlretrieve( url, "exchangeList.txt" )
#
# soup = bs( data, "html.parser" )
# exchange = soup.select( "td.tit a" )
# sale = soup.select( "td.sale" )
#
# for i in range( len( exchange ) ) :
#     print( exchange[i].string.strip(), "\t", sale[i].string.strip(),
#            sale[i].next_sibling.next_sibling.string.strip() )
    
# 네이버 실시간 뉴스
# url = "https://news.daum.net/"
# data = req.urlopen( url ).read().decode( "utf-8" )
# soup = bs( data, "html.parser" )
#
# import time
# titles = soup.select( "strong.tit_g a" )
# for i, title in enumerate( titles ) :
#     if i > 4 :
#         break
#     print( "<<", title.string.strip(), ">>" )
#
#     article_url = title.attrs["href"]
#     # print( article_url )
#     article_data = req.urlopen( article_url ).read().decode( "utf-8" )
#     # print( article_data )
#     soup = bs( article_data, "html.parser" )
#
#     ps = soup.select( "div#harmonyContainer section > p" )
#
#     for p in ps :
#         if p.string != None :
#             print( p.string )    
#
#     time.sleep( 3 )
#     print()
    
import requests
# req = requests.get( "http://www.google.co.kr/robots.txt" )
# text = req.text
# # print( text )    
# bin = req.content
# print( bin )
    
# req = requests.get( "https://img3.daumcdn.net/thumb/R658x0.q70/?fname=https://t1.daumcdn.net/news/202208/22/joongang/20220822094309182vvxf.jpg" )  
# with open( "test.jpg", "wb" ) as f:
#     f.write( req.content )
          
    
# 한빛미디어 로그인
# url = "https://www.hanbit.co.kr/member/login_proc.php"
# login_info = {
#     "m_id" : "filmal",
#     "m_passwd" : "abcd1234"
#     }
#
# session = requests.session()
# con = session.post( url, login_info )
# con.raise_for_status()
#
# url = "https://www.hanbit.co.kr/myhanbit/myhanbit.html"
# data = session.get( url )
# # print( data )
# data.raise_for_status()
#
# soup = bs( data.text, "html.parser" )
# mileage_title = soup.select_one( "dl.mileage_section1 dt" )
# mileage = soup.select_one( "dl.mileage_section1 dd span" )
# print( mileage_title.string, ":", mileage.string, "점" )
#
# ecoin_title = soup.select_one( "dl.mileage_section2 dt" )
# ecoin = soup.select_one( "dl.mileage_section2 dd span" )
# print( ecoin_title.string, ":", ecoin.string )
#
# trs = soup.select( "div.sm_myorder tr" )
# # print( trs )
# for i in range( 1, len(trs) ) :
#     print( trs[i].td.string, "\t", trs[i].a.string, "\t", trs[i].span.string )
    
    
    
    
# selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
# chrome_options = webdriver.ChromeOptions()
# driver = webdriver.Chrome( service=Service( ChromeDriverManager().install() ), 
#                            options=chrome_options )
# url = "http://naver.com"
# driver.get( url )
# time.sleep( 5 )
# driver.save_screenshot( "naver.png" )
# driver.quit()    
    
# 구글 로그인
from webdriver_manager.core.utils import ChromeType
# driver = webdriver.Chrome( service=Service( ChromeDriverManager( chrome_type=ChromeType.BRAVE ).install() ) )    
#
# url = "https://accounts.google.com/signin/v2/identifier"
# driver.implicitly_wait( 3 )   
# driver.get( url )
#
# driver.find_element( "id", "identifierId" ).send_keys( "instructor.set" )
# idbutton = driver.find_element( "id", "identifierNext" )
# time.sleep( 2 )
# idbutton.click()
#
# time.sleep( 5 )
 
# 한빛미디어 로그인
# url = "https://www.hanbit.co.kr/member/login.html"
# driver = webdriver.Chrome( service=Service( ChromeDriverManager( chrome_type=ChromeType.BRAVE ).install() ) )
# driver.implicitly_wait( 3 )
# driver.get( url )
# driver.find_element( "id", "m_id" ).send_keys( "filmal" )
# driver.find_element( "id", "m_passwd" ).send_keys( "abcd1234" )
# time.sleep( 2 )
# loginbutton = driver.find_element( "id", "login_btn" )
# loginbutton.click()
# time.sleep( 5 )    
    
    
# 공공데이터 포털 - 국토교통부_아파트매매 실거래 상세 자료
import urllib.request as req
import urllib.parse as pa
url = "http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev?LAWD_CD=11545&DEAL_YMD=202207&serviceKey=OUwSlot%2BeCADq1M9zzdj8Sh1Ni9C4Iiaj9VqSEnyvikodjynkoS1hrbUsP6mSENccvTJH%2FDe3s3y7i836Lk7ew%3D%3D"
# url = "http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev"
# values = {
#     "LAWD_CD" : 11545,
#     "DEAL_YMD" : 202207,
#     "serviceKey" : "OUwSlot%2BeCADq1M9zzdj8Sh1Ni9C4Iiaj9VqSEnyvikodjynkoS1hrbUsP6mSENccvTJH%2FDe3s3y7i836Lk7ew%3D%3D"
#     }
#
# params = pa.urlencode( values )
# url = url + "?" + params
# print( url )

# data = req.urlopen( url ).read().decode( "utf-8" )
# # print( data )
# with open( "apart202207.txt", "w", encoding="utf-8" ) as f :
#     f.write( data )    
# data = data.replace( "거래금액", "price" )
# data = data.replace( "법정동", "dong" )
# data = data.replace( "아파트", "apartname" )
# data = data.replace( "월", "month" )
# data = data.replace( "일", "day" )
#
# soup = bs( data, "html.parser" )
# items = soup.select( "item" )
# for item in items :
#     # print( item )
#     print( item.dong.string, item.month.string, "-", item.day.string,
#            item.apartname.string, item.price.string )

    
# JSON 데이터
# import json
# import os.path
# url = "https://api.github.com/repositories"
# filename = "github.txt"
#
# if not os.path.exists( filename ) :
#     req.urlretrieve( url, filename )
# items = json.load( open( filename, "rt", encoding="utf-8" ) )    
#
# for item in items :
#     print( item["id"], item["name"], item["owner"]["login"] )
    
# CSV 데이터
# import csv, codecs
# filename = "test.csv"
# write = codecs.open( filename, "w", encoding="utf-8" )    
# writer = csv.writer( write, delimiter="," )
# writer.writerow( ["아이디","이름","가격"] )
# writer.writerow( ["1000","HDD",200000] )
# writer.writerow( ["1001","SSD",300000] )
# writer.writerow( ["1002","Monitor",150000] )
# writer.writerow( ["1003","Mouse",20000] )
# writer.writerow( ["1004","Keyboard",30000] )
# print( "파일 생성 완료" )
# write.close()
#
# readcsv = codecs.open( filename, "r", encoding="utf-8" ).read()
# data = []
# rows = readcsv.split( "\r\n" )
# # print( rows )
# for row in rows :
#     if row == "" :
#         break
#     cells = row.split( "," )
#     data.append( cells )
#
# print( data )
# for d in data :
#     print( d[0], "\t", d[1], "\t", d[2] )
    
# 엑셀 데이터
import openpyxl
filename = "stat_100701.xlsx"    
wb = openpyxl.load_workbook( filename )
# ws = wb.worksheets[0]
ws = wb.active
# for row in ws.rows :
#     for data in row :
#         if data.value == None :
#             print( "", end="\t" )
#         else :   
#             print( data.value, end="\t" )
#     print()

data = []
for row in ws.rows :
    if row[9] != None and row[10] != None :
        data.append( [row[9].value, row[10].value] )
del( data[0:4])  

data = sorted( data, key=lambda x : x[1], reverse=True )
for d in data :      
    print( d )         
    
savefile = "population.xlsx"
swb = openpyxl.Workbook()
sws = swb.active
# sws = swb.create_sheet( title="인구" )
for i, d in enumerate( data ) :
    sws.cell( row=i+1, column=1, value=d[0] )   
    sws.cell( row=i+1, column=2, value=d[1] ) 
swb.save( savefile )
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    

